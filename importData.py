import pyodbc
import os
import openpyxl
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import absolute_coordinate
##connecting database
conn = pyodbc.connect(
    r"Driver={ODBC Driver 17 for SQL Server};"
    r"Server=DESKTOP-US471NN\MSSQLSERVER01;"
    r"Database=PrismDB;"
    r"trusted_connection=yes;"
)
cursor = conn.cursor()
book = load_workbook('QJS Daily Sales Report  30-Mar-2023.xlsx')
sheet_names = sorted(book.sheetnames)
dates = ['2023-03-14', '2023-03-15', '2023-03-16', '2023-03-17', '2023-03-18', '2023-03-19', '2023-03-20', '2023-03-21', '2023-03-22', '2023-03-23', '2023-03-24', '2023-03-25', '2023-03-26', '2023-03-27', '2023-03-28', '2023-03-29', '2023-03-30', '2023-03-31', '2023-04-01', '2023-04-02', '2023-04-03', '2023-04-04', '2023-04-05', '2023-04-06', '2023-04-07', '2023-04-08', '2023-04-09', '2023-04-10', '2023-04-11', '2023-04-12', '2023-04-13', '2023-04-14']
columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'S', 'T', 'U']
for name in sheet_names:
    sheet = book[name]
    city = ''
    store = ''
    area = ''
    BA = ''
    if isinstance(sheet['C2'].value, str):
        if sheet['C2'].value.startswith("City:"):
            city = sheet['C2'].value[7:]
    
            if isinstance(sheet['C3'].value, str):
                if sheet['C3'].value.startswith("Store:"):
                    dash = (sheet['C3'].value.rfind("-"))
                    if dash > 0:
                        store = (sheet['C3'].value[7:dash-1]).strip()
            
            if isinstance(sheet['C3'].value, str):
                if sheet['C3'].value.startswith("Store:"):
                    dash = (sheet['C3'].value.rfind("-"))+1
                    if dash > 0:
                        area = (sheet['C3'].value[dash:]).strip()
            
            if isinstance(sheet['C4'].value, str):
                if sheet['C4'].value.startswith("BA"):
                    BA = (sheet['C4'].value[9:]).strip()
            for row in range (8, 15):
                dateCur = 0
                for column in columns:
                    date = dates[dateCur]
                    try:
                        date_obj = datetime.datetime.strptime(date, '%Y-%m-%d').date()
                    except:
                        date_obj = datetime.date(2023, 1, 5)
                    variant = 0
                    brand = 0
                    size = 0
                    subproduct = 0
                    totalSales = 0
                    cellVal = sheet[column + str(row)].value
                    if row < 11 :
                        variant  = 1
                        brand = 1
                        if row == 8:
                            size = 1
                            subproduct = 1
                        if row == 9:
                            size = 2
                            subproduct = 1
                        if row == 10:
                            size = 3
                            subproduct = 1
                   
                    else:
                        variant  = 2
                        if row == 11:
                            size = 1
                            subproduct = 2
                            brand = 1
                        
                        if row == 12:
                            size = 1
                            subproduct = 1
                            brand = 2
                        if row == 13:
                            size = 1
                            subproduct = 1
                            brand = 3
                        if row == 14:
                            size = 1
                            subproduct = 1
                            brand = 4
                    totalSales = cellVal
                    if totalSales is not None:
                        cursor.execute("SELECT id FROM city WHERE name = ?;", (city,))
                        ans = cursor.fetchone()
                        if ans:
                            cityId = ans[0]
                        else:
                            cursor.execute("select max(id) from city;")
                            ans = cursor.fetchone()
                            if ans[0] is not None:   
                                cityId = ans[0] + 1
                            else: 
                                cityId = 1
                            cursor.execute("insert into city(id, name) values(?,?);", cityId, city)
                        cursor.commit()

                        cursor.execute("SELECT id FROM storeNames WHERE name = ?;", (store,))
                        ans = cursor.fetchone()
                        if ans:
                            storeId = ans[0]
                        else:
                            cursor.execute("select max(id) from storeNames;")
                            ans = cursor.fetchone()
                            if ans[0] is not None:   
                                storeId = ans[0] + 1
                            else: 
                                storeId = 1
                            cursor.execute("insert into storeNames(id, name) values(?,?);", storeId, store)
                        cursor.commit()

                        cursor.execute("SELECT id FROM area WHERE name = ?;", (area,))
                        ans = cursor.fetchone()
                        if ans:
                            areaId = ans[0]
                        else:
                            cursor.execute("select max(id) from area;")
                            ans = cursor.fetchone()
                            if ans[0] is not None:   
                                areaId = ans[0] + 1
                            else: 
                                areaId = 1
                            cursor.execute("insert into area(name) values(?);", area)
                        cursor.commit()

                        cursor.execute("SELECT id FROM BA WHERE name = ?;", (BA,))
                        ans = cursor.fetchone()
                        if ans:
                            BAId = ans[0]
                        else:
                            cursor.execute("select max(id) from BA;")
                            ans = cursor.fetchone()
                            if ans[0] is not None:   
                                BAId = ans[0] + 1
                            else: 
                                BAId = 1
                            cursor.execute("insert into BA(id, name) values(?,?);", BAId, BA)
                        cursor.commit()
                        cursor.execute("insert into masterTable(date, city, store, area, BA, variant, brand, size, subproduct, totalSales) values(?,?,?,?,?,?,?,?,?,?);", date_obj, cityId, storeId, areaId, BAId, variant, brand, size, subproduct, totalSales)
                        dateCur = dateCur + 1
                        cursor.commit()
